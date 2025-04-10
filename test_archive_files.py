import os
import csv
import zipfile
from io import TextIOWrapper
from openpyxl import load_workbook
from pypdf import PdfReader

archive_path = os.path.join("files", "archive.zip")


def test_csv_file():
    with zipfile.ZipFile(archive_path) as zf:
        with zf.open('Test_CSV.csv') as file:
            reader = list(csv.reader(TextIOWrapper(file, 'utf-8-sig'), delimiter=';'))
            assert reader == [['a1', 'b1', 'c1'],
                              ['a2', 'b2', 'c2'],
                              ['a3', 'b3', 'c3']]


def test_xlsx_file():
    with zipfile.ZipFile(archive_path) as zf:
        with zf.open('Test_Excel.xlsx') as file:
            sheet = load_workbook(file).active
            values = [cell.value for row in sheet.iter_rows() for cell in row]
            assert values == ['a1', 'b1', 'c1',
                              'a2', 'b2', 'c2',
                              'a3', 'b3', 'c3']


def test_pdf_file():
    with zipfile.ZipFile(archive_path) as zf:
        with zf.open('Test_Pdf.pdf') as file:
            reader = PdfReader(file)
            text = reader.pages[0].extract_text()
            assert 'This is a test PDF document' in text
